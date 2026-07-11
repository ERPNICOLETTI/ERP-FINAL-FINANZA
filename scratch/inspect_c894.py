import openpyxl
from openpyxl.styles import PatternFill

target_file = r"C:\Users\essao\Desktop\Escritorio\gestion_quimica\uploads\N.P._PATAGONIA_25-11-2025_CON_OFERTAS_DE_DICIEMBRE.xlsx"
wb = openpyxl.load_workbook(target_file, data_only=True)
ws = wb['LISTA Y OFERTAS NOV 2025']

cell = ws['C894']
print("--- Celda C894 ---")
print("Valor:", cell.value)
print("Tipo de Fill:", type(cell.fill))
print("Fill type:", cell.fill.fill_type)
if cell.fill and hasattr(cell.fill, 'start_color'):
    print("Start color:", cell.fill.start_color)
    print("Start color RGB:", getattr(cell.fill.start_color, 'rgb', None))
    print("Start color Indexed:", getattr(cell.fill.start_color, 'indexed', None))
    print("Start color Theme:", getattr(cell.fill.start_color, 'theme', None))
    print("Start color Tint:", getattr(cell.fill.start_color, 'tint', None))

if cell.font:
    print("Font color:", cell.font.color)
    if cell.font.color:
        print("Font color RGB:", getattr(cell.font.color, 'rgb', None))
        print("Font color Theme:", getattr(cell.font.color, 'theme', None))

print("\n--- Verificando si está combinada ---")
for merged_range in ws.merged_cells.ranges:
    if 'C894' in merged_range:
        print("Combinada en rango:", merged_range)

wb.close()
