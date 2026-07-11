import openpyxl

target_file = r"C:\Users\essao\Desktop\Escritorio\gestion_quimica\uploads\N.P._PATAGONIA_25-11-2025_CON_OFERTAS_DE_DICIEMBRE.xlsx"

# Abrir con data_only=False para ver la fórmula en texto
wb = openpyxl.load_workbook(target_file, data_only=False)
ws = wb['OFERTAS DE STOCK NOV 2025']

cell = ws['E32']
print("Fórmula en E32:", cell.value)

# Abrir con data_only=True para ver el valor resultante
wb_val = openpyxl.load_workbook(target_file, data_only=True)
ws_val = wb_val['OFERTAS DE STOCK NOV 2025']
print("Valor evaluado en E32:", ws_val['E32'].value)

wb.close()
wb_val.close()
