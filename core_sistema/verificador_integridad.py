import os
import hashlib
import sqlite3
import openpyxl
import csv

# SISTEMA INTEGRADO DE AUDITORÍA Y VERIFICACIÓN CRUZADA v1.3.0 🛡️🧮

class VerificadorIntegridadERP:
    def __init__(self, db_path: str = 'c:/Users/essao/Desktop/ERP FINAL/erp_nicoletti.db'):
        self.db_path = db_path

    def verificar_ingesta_banco_excel(self, filepath: str, staging_id: int) -> dict:
        """
        Realiza una reconciliación matemática 1:1 para Galicia (Excel).
        """
        if not os.path.exists(filepath):
            return {"status": "ERROR", "mensaje": f"Archivo físico no encontrado: {filepath}"}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb['Cuentas']
            excel_filas = 0
            excel_debitos = 0.0
            excel_creditos = 0.0

            for r in range(7, ws.max_row + 1):
                f_val = ws.cell(r, 1).value
                m_val = ws.cell(r, 2).value
                deb_val = ws.cell(r, 3).value
                cred_val = ws.cell(r, 4).value
                
                if f_val and m_val:
                    excel_filas += 1
                    
                    def to_float(val):
                        if not val: return 0.0
                        try:
                            return float(str(val).replace('.', '').replace(',', '.'))
                        except:
                            return 0.0
                            
                    excel_debitos += abs(to_float(deb_val))
                    excel_creditos += abs(to_float(cred_val))
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error al leer Excel Galicia: {e}"}

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            st = conn.execute("SELECT id, hash_sha256 FROM core_staging_raw WHERE id = ?", (staging_id,)).fetchone()
            if not st:
                conn.close()
                return {"status": "ERROR", "mensaje": f"No se encontró el registro Staging ID {staging_id}"}
                
            bm = conn.execute("""
                SELECT count(*) as count, 
                       sum(case when importe < 0 then abs(importe) else 0 end) as debitos, 
                       sum(case when importe > 0 then importe else 0 end) as creditos 
                FROM bancos_movimientos 
                WHERE raw_ingesta_id = ?
            """, (staging_id,)).fetchone()
            conn.close()
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error de base de datos: {e}"}

        db_filas = bm['count'] or 0
        db_debitos = bm['debitos'] or 0.0
        db_creditos = bm['creditos'] or 0.0

        dif_filas = excel_filas - db_filas
        dif_debitos = abs(excel_debitos - db_debitos)
        dif_creditos = abs(excel_creditos - db_creditos)

        conciliado = (dif_filas == 0 and dif_debitos < 0.01 and dif_creditos < 0.01)

        return {
            "status": "OK" if conciliado else "DISCREPANCIA",
            "hash_sha256": st['hash_sha256'],
            "origen": {"filas": excel_filas, "debitos": round(excel_debitos, 2), "creditos": round(excel_creditos, 2)},
            "destino": {"filas": db_filas, "debitos": round(db_debitos, 2), "creditos": round(db_creditos, 2)},
            "diferencias": {"filas": dif_filas, "debitos": round(dif_debitos, 2), "creditos": round(dif_creditos, 2)}
        }

    def verificar_ingesta_banco_chubut(self, filepath: str, staging_id: int) -> dict:
        """
        Realiza una reconciliación matemática 1:1 para Banco Chubut (Excel).
        """
        if not os.path.exists(filepath):
            return {"status": "ERROR", "mensaje": f"Archivo físico no encontrado: {filepath}"}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb['Historicos']
            excel_filas = 0
            excel_total = 0.0

            for r in range(11, ws.max_row + 1):
                f = ws.cell(r, 2).value
                m = ws.cell(r, 3).value
                imp = ws.cell(r, 5).value
                if f and m:
                    excel_filas += 1
                    excel_total += float(str(imp or 0.0))
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error al leer Excel Chubut: {e}"}

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            st = conn.execute("SELECT id, hash_sha256 FROM core_staging_raw WHERE id = ?", (staging_id,)).fetchone()
            if not st:
                conn.close()
                return {"status": "ERROR", "mensaje": f"No se encontró el registro Staging ID {staging_id}"}
                
            db = conn.execute("SELECT count(*) as count, sum(importe) as total FROM bancos_movimientos WHERE raw_ingesta_id = ?", (staging_id,)).fetchone()
            conn.close()
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error de base de datos: {e}"}

        db_filas = db['count'] or 0
        db_total = db['total'] or 0.0

        dif_filas = excel_filas - db_filas
        dif_total = abs(excel_total - db_total)

        conciliado = (dif_filas == 0 and dif_total < 0.01)

        return {
            "status": "OK" if conciliado else "DISCREPANCIA",
            "hash_sha256": st['hash_sha256'],
            "origen": {"filas": excel_filas, "total": round(excel_total, 2)},
            "destino": {"filas": db_filas, "total": round(db_total, 2)},
            "diferencias": {"filas": dif_filas, "total": round(dif_total, 2)}
        }

    def verificar_ingesta_banco_hipotecario(self, filepath: str, staging_id: int) -> dict:
        """
        Realiza una reconciliación matemática 1:1 para Banco Hipotecario (Excel Pesos y USD).
        """
        if not os.path.exists(filepath):
            return {"status": "ERROR", "mensaje": f"Archivo físico no encontrado: {filepath}"}

        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            ws = wb.active
            excel_filas = 0
            excel_total = 0.0

            for r in range(6, ws.max_row + 1):
                f = ws.cell(r, 1).value
                m = ws.cell(r, 2).value
                imp = ws.cell(r, 3).value
                if f and m:
                    excel_filas += 1
                    
                    def to_float(val):
                        if not val: return 0.0
                        try:
                            return float(str(val).replace('.', '').replace(',', '.'))
                        except:
                            return 0.0
                            
                    excel_total += to_float(imp)
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error al leer Excel Hipotecario: {e}"}

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            st = conn.execute("SELECT id, hash_sha256 FROM core_staging_raw WHERE id = ?", (staging_id,)).fetchone()
            if not st:
                conn.close()
                return {"status": "ERROR", "mensaje": f"No se encontró el registro Staging ID {staging_id}"}
                
            db = conn.execute("SELECT count(*) as count, sum(importe) as total FROM bancos_movimientos WHERE raw_ingesta_id = ?", (staging_id,)).fetchone()
            conn.close()
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error de base de datos: {e}"}

        db_filas = db['count'] or 0
        db_total = db['total'] or 0.0

        dif_filas = excel_filas - db_filas
        dif_total = abs(excel_total - db_total)

        conciliado = (dif_filas == 0 and dif_total < 0.01)

        return {
            "status": "OK" if conciliado else "DISCREPANCIA",
            "hash_sha256": st['hash_sha256'],
            "origen": {"filas": excel_filas, "total": round(excel_total, 2)},
            "destino": {"filas": db_filas, "total": round(db_total, 2)},
            "diferencias": {"filas": dif_filas, "total": round(dif_total, 2)}
        }

    def verificar_ingesta_mercadopago(self, filepath: str, staging_id: int) -> dict:
        """
        Realiza una reconciliación matemática de doble capa 1:1 para Mercado Pago (CSV).
        """
        if not os.path.exists(filepath):
            return {"status": "ERROR", "mensaje": f"Archivo físico no encontrado: {filepath}"}

        try:
            csv_filas = 0
            csv_neto = 0.0
            with open(filepath, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                headers = next(reader)
                idx_neto = headers.index('SETTLEMENT_NET_AMOUNT')
                for row in reader:
                    if row:
                        csv_filas += 1
                        csv_neto += float(row[idx_neto])
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error al leer CSV Mercado Pago: {e}"}

        try:
            conn = sqlite3.connect(self.db_path)
            conn.row_factory = sqlite3.Row
            st = conn.execute("SELECT id, hash_sha256 FROM core_staging_raw WHERE id = ?", (staging_id,)).fetchone()
            if not st:
                conn.close()
                return {"status": "ERROR", "mensaje": f"No se encontró el registro Staging ID {staging_id}"}
                
            bm = conn.execute("SELECT count(*) as count, sum(importe) as neto FROM bancos_movimientos WHERE raw_ingesta_id = ?", (staging_id,)).fetchone()
            conn.close()
        except Exception as e:
            return {"status": "ERROR", "mensaje": f"Error de base de datos: {e}"}

        db_filas = bm['count'] or 0
        db_neto = bm['neto'] or 0.0

        dif_filas = csv_filas - db_filas
        dif_neto = abs(csv_neto - db_neto)

        conciliado = (dif_filas == 0 and dif_neto < 0.01)

        return {
            "status": "OK" if conciliado else "DISCREPANCIA",
            "hash_sha256": st['hash_sha256'],
            "origen": {"filas": csv_filas, "total": round(csv_neto, 2)},
            "destino": {"filas": db_filas, "total": round(db_neto, 2)},
            "diferencias": {"filas": dif_filas, "total": round(dif_neto, 2)}
        }

    def verificar_y_reportar(self, filepath: str, staging_id: int) -> dict:
        """
        Detecta el tipo de archivo y despacha al verificador correspondiente.
        """
        filename = os.path.basename(filepath).upper()
        if "SETTLEMENT" in filename or filename.endswith(".CSV"):
            res = self.verificar_ingesta_mercadopago(filepath, staging_id)
            print(f"🛡️ [VERIFICADOR MP] Resultado: {res['status']} | Desvío Neto: {res.get('diferencias', {}).get('total', 0)}")
            return res
        elif "CHUBUT" in filename or "HISTORICOS" in filename:
            res = self.verificar_ingesta_banco_chubut(filepath, staging_id)
            print(f"🛡️ [VERIFICADOR CHUBUT] Resultado: {res['status']} | Desvío Neto: {res.get('diferencias', {}).get('total', 0)}")
            return res
        elif "HIPOTECARIO" in filename or "9087" in filename or "2646" in filename:
            res = self.verificar_ingesta_banco_hipotecario(filepath, staging_id)
            print(f"🛡️ [VERIFICADOR HIPOTECARIO] Resultado: {res['status']} | Desvío Neto: {res.get('diferencias', {}).get('total', 0)}")
            return res
        else:
            res = self.verificar_ingesta_banco_excel(filepath, staging_id)
            print(f"🛡️ [VERIFICADOR GALICIA] Resultado: {res['status']} | Desvío Neto: {res.get('diferencias', {}).get('debitos', 0) + res.get('diferencias', {}).get('creditos', 0)}")
            return res
