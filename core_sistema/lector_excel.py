import os
import re
import csv
import hashlib
from datetime import datetime

# Intentar importar openpyxl
try:
    import openpyxl
    from openpyxl.utils import range_boundaries
except ImportError:
    openpyxl = None

try:
    import xlrd
except ImportError:
    xlrd = None


def calcular_hash_archivo(filepath: str) -> str:
    """Calcula el hash SHA-256 de un archivo físico para control de duplicados/linaje."""
    sha256 = hashlib.sha256()
    with open(filepath, 'rb') as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            sha256.update(chunk)
    return sha256.hexdigest()


class LectorExcelUniversal:
    """
    Parser general de Excel (.xlsx, .xls) y CSV.
    Extrae la matriz física de celdas, valores, colores de fondo/fuente
    y metadatos estructurales directamente en memoria.
    """

    def extraer_raw(self, filepath: str) -> dict:
        """
        Interroga el archivo y extrae metadatos globales, pestañas y el ADN completo
        de cada celda (fuentes, bordes, alineación, fórmulas, celdas combinadas, etc.).
        """
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"El archivo no existe: {filepath}")

        ext = os.path.splitext(filepath)[1].lower()
        pestanas_info = []
        filas_extraidas = []

        if ext == '.csv':
            nombre_pestana = "CSV_DEFAULT"
            pestanas_info.append(nombre_pestana)
            with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for r_idx, row in enumerate(reader):
                    for c_idx, val in enumerate(row):
                        col_letter = self._convertir_columna_a_letra(c_idx + 1)
                        coord = f"{col_letter}{r_idx + 1}"
                        filas_extraidas.append({
                            "pestana": nombre_pestana,
                            "coordenada": coord,
                            "valor": val,
                            "color": None,
                            "color_fuente": None,
                            "metadatos": {}
                        })

        elif ext == '.xlsx':
            if not openpyxl:
                raise ImportError("La librería openpyxl no está instalada.")
            
            # Abrir en modo fórmulas y modo valores
            wb_formulas = openpyxl.load_workbook(filepath, data_only=False)
            wb_values = openpyxl.load_workbook(filepath, data_only=True)
            pestanas_info = wb_values.sheetnames
            
            for sheet_name in pestanas_info:
                ws_val = wb_values[sheet_name]
                ws_form = wb_formulas[sheet_name]
                
                merged_ranges = ws_val.merged_cells.ranges
                
                autofiltro_ref = ws_val.auto_filter.ref
                autofiltro_col_start, autofiltro_row_start = 0, 0
                autofiltro_col_end, autofiltro_row_end = 0, 0
                if autofiltro_ref:
                    min_c, min_r, max_c, max_r = range_boundaries(autofiltro_ref)
                    autofiltro_col_start, autofiltro_row_start = min_c, min_r
                    autofiltro_col_end, autofiltro_row_end = max_c, max_r
                
                for r_idx, row_val in enumerate(ws_val.iter_rows()):
                    row_idx_num = r_idx + 1
                    for c_idx, cell_val in enumerate(row_val):
                        col_idx_num = c_idx + 1
                        coord = cell_val.coordinate
                        cell_form = ws_form[coord]
                        
                        val = cell_val.value
                        formula = None
                        if cell_form.value and isinstance(cell_form.value, str) and cell_form.value.startswith('='):
                            formula = cell_form.value

                        # Color fondo
                        color_hex = None
                        if cell_val.fill and getattr(cell_val.fill, 'fill_type', None) is not None:
                            start_color = getattr(cell_val.fill, 'start_color', None)
                            if start_color:
                                rgb = getattr(start_color, 'rgb', None)
                                if isinstance(rgb, str) and rgb != "00000000" and len(rgb) >= 6:
                                    if len(rgb) == 8 and rgb.startswith('FF'):
                                        rgb = rgb[2:]
                                    color_hex = f"#{rgb}" if not rgb.startswith('#') else rgb

                        # Color fuente
                        color_fuente_hex = None
                        if cell_val.font and getattr(cell_val.font, 'color', None) is not None:
                            font_color = cell_val.font.color
                            rgb = getattr(font_color, 'rgb', None)
                            if isinstance(rgb, str) and rgb != "00000000" and len(rgb) >= 6:
                                if len(rgb) == 8 and rgb.startswith('FF'):
                                    rgb = rgb[2:]
                                color_fuente_hex = f"#{rgb}" if not rgb.startswith('#') else rgb

                        # Metadatos de estilo
                        meta = {
                            "fuente": {
                                "nombre": getattr(cell_val.font, 'name', None),
                                "tamano": getattr(cell_val.font, 'size', None),
                                "negrita": getattr(cell_val.font, 'bold', False),
                                "cursiva": getattr(cell_val.font, 'italic', False),
                                "subrayado": getattr(cell_val.font, 'underline', None),
                                "tachar": getattr(cell_val.font, 'strike', False),
                                "color": color_fuente_hex
                            },
                            "relleno": {
                                "tipo": getattr(cell_val.fill, 'fill_type', None),
                                "color_fondo": color_hex
                            },
                            "borde": {
                                "arriba": getattr(getattr(cell_val.border, 'top', None), 'style', None),
                                "abajo": getattr(getattr(cell_val.border, 'bottom', None), 'style', None),
                                "izquierda": getattr(getattr(cell_val.border, 'left', None), 'style', None),
                                "derecha": getattr(getattr(cell_val.border, 'right', None), 'style', None)
                            },
                            "alineacion": {
                                "horizontal": getattr(cell_val.alignment, 'horizontal', None),
                                "vertical": getattr(cell_val.alignment, 'vertical', None),
                                "ajustar_texto": getattr(cell_val.alignment, 'wrap_text', None)
                            },
                            "estructura": {
                                "formula": formula,
                                "es_combinada": any(coord in rng for rng in merged_ranges),
                                "tiene_filtro": (autofiltro_row_start == row_idx_num and autofiltro_col_start <= col_idx_num <= autofiltro_col_end)
                            },
                            "formato_numero": cell_val.number_format,
                            "dimensiones": {
                                "ancho_columna": ws_val.column_dimensions[cell_val.column_letter].width,
                                "alto_fila": ws_val.row_dimensions[row_idx_num].height
                            }
                        }

                        filas_extraidas.append({
                            "pestana": sheet_name,
                            "coordenada": coord,
                            "valor": val,
                            "color": color_hex,
                            "color_fuente": color_fuente_hex,
                            "metadatos": meta
                        })
            wb_values.close()
            wb_formulas.close()

        elif ext == '.xls':
            if not xlrd:
                raise ImportError("La librería xlrd no está instalada.")
            wb = xlrd.open_workbook(filepath, formatting_info=True)
            pestanas_info = wb.sheet_names()
            for sheet_idx, sheet_name in enumerate(pestanas_info):
                ws = wb.sheet_by_index(sheet_idx)
                for r_idx in range(ws.nrows):
                    for c_idx in range(ws.ncols):
                        val = ws.cell_value(r_idx, c_idx)
                        xf_idx = ws.cell_xf_index(r_idx, c_idx)
                        xf = wb.xf_list[xf_idx]
                        bg_color_hex = None
                        font_color_hex = None
                        
                        meta = {}
                        if xf:
                            if xf.background:
                                pattern_colour_index = xf.background.pattern_colour_index
                                if pattern_colour_index in wb.colour_map:
                                    rgb = wb.colour_map[pattern_colour_index]
                                    bg_color_hex = f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                            if xf.font_index < len(wb.font_list):
                                font = wb.font_list[xf.font_index]
                                if font and font.colour_index in wb.colour_map:
                                    rgb = wb.colour_map[font.colour_index]
                                    font_color_hex = f"#{rgb[0]:02X}{rgb[1]:02X}{rgb[2]:02X}"
                                    meta = {
                                        "fuente": {
                                            "nombre": font.name,
                                            "tamano": font.height / 20.0,
                                            "negrita": font.bold,
                                            "cursiva": font.italic,
                                            "underline": font.underline,
                                            "color": font_color_hex
                                        }
                                    }
                        
                        col_letter = self._convertir_columna_a_letra(c_idx + 1)
                        coord = f"{col_letter}{r_idx + 1}"
                        filas_extraidas.append({
                            "pestana": sheet_name,
                            "coordenada": coord,
                            "valor": val,
                            "color": bg_color_hex,
                            "color_fuente": font_color_hex,
                            "metadatos": meta
                        })
        else:
            raise ValueError(f"Formato no soportado: {ext}")

        return {
            "nombre_archivo": os.path.basename(filepath),
            "hash_sha256": calcular_hash_archivo(filepath),
            "cantidad_pestanas": len(pestanas_info),
            "pestanas": pestanas_info,
            "filas": filas_extraidas
        }

    def _convertir_columna_a_letra(self, col_idx: int) -> str:
        result = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            result = chr(65 + remainder) + result
        return result

    def transformar_celda(self, valor, es_importe=False, es_identificador=False, es_fecha=False):
        """Aplica reglas de tipado y saneamiento a un valor individual."""
        if valor is None or str(valor).strip() == "":
            return None
        val_str = str(valor).strip()
        if es_importe:
            try:
                clean_val = val_str.replace("$", "").replace(" ", "")
                if "," in clean_val and "." in clean_val:
                    if clean_val.find(",") < clean_val.find("."):
                        clean_val = clean_val.replace(",", "")
                    else:
                        clean_val = clean_val.replace(".", "").replace(",", ".")
                elif "," in clean_val:
                    clean_val = clean_val.replace(",", ".")
                return int(round(float(clean_val) * 100))
            except ValueError:
                raise ValueError(f"No se pudo convertir '{valor}' a Importe.")
        if es_identificador:
            return val_str.upper()
        if es_fecha:
            pattern_dmy = r'^(\d{1,2})[/-](\d{1,2})[/-](\d{4})'
            pattern_ymd = r'^(\d{4})[/-](\d{1,2})[/-](\d{1,2})'
            m_dmy = re.match(pattern_dmy, val_str)
            if m_dmy:
                return f"{m_dmy.group(3)}-{m_dmy.group(2).zfill(2)}-{m_dmy.group(1).zfill(2)}"
            m_ymd = re.match(pattern_ymd, val_str)
            if m_ymd:
                return f"{m_ymd.group(1)}-{m_ymd.group(2).zfill(2)}-{m_ymd.group(3).zfill(2)}"
            try:
                return datetime.strptime(val_str, "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
            except ValueError:
                pass
            raise ValueError(f"Fecha inválida: '{valor}'")
        return valor
