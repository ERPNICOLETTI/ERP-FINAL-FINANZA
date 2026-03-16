
import openpyxl
import os

def count_invoices(file_path):
    print(f"\nAnalizando: {os.path.basename(file_path)}")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        total_valid = 0
        all_numbers = set()
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.values)
            if not rows: continue
            
            header_row_idx = -1
            col_map = {}
            
            # 1. Encontrar cabecera
            for i, row in enumerate(rows[:50]):
                row_str = " ".join([str(c).lower() for c in row if c]).replace("None", "")
                if ("proveedor" in row_str or "denominaci" in row_str or "nombre" in row_str) and \
                   ("iva" in row_str or "neto" in row_str or "total" in row_str or "monto" in row_str):
                    header_row_idx = i
                    # 2. Mapear columnas
                    for idx, cell in enumerate(row):
                        h = str(cell or "").lower().strip()
                        if "punto de venta" in h or "p.v." in h: col_map['pv'] = idx
                        if "número desde" in h or "numero desde" in h or "nro. desde" in h: col_map['num'] = idx
                        if any(x in h for x in ["numero", "número", "nro", "factura", "comprobante"]):
                            if 'numero_calim' not in col_map: col_map['numero_calim'] = idx
                    break
            
            if header_row_idx == -1:
                print(f"  - Hoja '{sheet_name}': No se detectaron encabezados de factura.")
                continue

            # 3. Contar facturas
            valid_in_sheet = 0
            for row in rows[header_row_idx + 1:]:
                if not any(row): continue
                
                num_comp = ""
                # Lógica del ERP
                if 'numero_calim' in col_map and row[col_map['numero_calim']]:
                    num_comp = "".join(filter(str.isdigit, str(row[col_map['numero_calim']])))
                elif 'pv' in col_map and 'num' in col_map and row[col_map['pv']] and row[col_map['num']]:
                    pv = "".join(filter(str.isdigit, str(row[col_map['pv']]))).zfill(4)
                    num = "".join(filter(str.isdigit, str(row[col_map['num']]))).zfill(8)
                    num_comp = pv + num
                
                num_comp = num_comp.lstrip('0')
                if len(num_comp) >= 2:
                    valid_in_sheet += 1
                    all_numbers.add(num_comp)
            
            print(f"  - Hoja '{sheet_name}': {valid_in_sheet} registros encontrados.")
            total_valid += valid_in_sheet
            
        return len(all_numbers), total_valid
    except Exception as e:
        print(f"Error leyendo {file_path}: {e}")
        return 0, 0

paths = [
    r"C:\Users\essao\Downloads\Facturas de Compra (7).xlsx",
    r"C:\Users\essao\Downloads\Mis Comprobantes Recibidos - CUIT 27329549971 (5).xlsx"
]

grand_total_unique = set()
for p in paths:
    unique, total = count_invoices(p)
    # Re-analizar para set global si fuera necesario, pero count_invoices ya lo hace por archivo
    # Para ser exacto, re-corremos para llenar grand_total_unique
    try:
        wb = openpyxl.load_workbook(p, data_only=True)
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.values)
            header_row_idx = -1
            col_map = {}
            for i, row in enumerate(rows[:50]):
                row_str = " ".join([str(c).lower() for c in row if c])
                if ("proveedor" in row_str or "denominaci" in row_str) and ("iva" in row_str or "total" in row_str):
                    header_row_idx = i
                    for idx, cell in enumerate(row):
                        h = str(cell or "").lower().strip()
                        if "punto de venta" in h: col_map['pv'] = idx
                        if "número desde" in h: col_map['num'] = idx
                        if any(x in h for x in ["numero", "nro", "factura"]):
                            if 'numero_calim' not in col_map: col_map['numero_calim'] = idx
                    break
            if header_row_idx != -1:
                for row in rows[header_row_idx + 1:]:
                    num_comp = ""
                    if 'numero_calim' in col_map and row[col_map['numero_calim']]:
                        num_comp = "".join(filter(str.isdigit, str(row[col_map['numero_calim']])))
                    elif 'pv' in col_map and 'num' in col_map and row[col_map['pv']] and row[col_map['num']]:
                        pv = "".join(filter(str.isdigit, str(row[col_map['pv']]))).zfill(4)
                        num = "".join(filter(str.isdigit, str(row[col_map['num']]))).zfill(8)
                        num_comp = pv + num
                    num_comp = num_comp.lstrip('0')
                    if len(num_comp) >= 2: grand_total_unique.add(num_comp)
    except: pass

print(f"\n--- RESUMEN FINAL ---")
print(f"Total de facturas únicas encontradas (sin repetir): {len(grand_total_unique)}")
