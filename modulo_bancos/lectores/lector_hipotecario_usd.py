import os
import hashlib
import openpyxl
from datetime import datetime
from modulo_bancos import storage_bancos

def procesar_archivo(filepath: str, reconstruir: bool = False) -> tuple[bool, dict]:
    """
    Procesador ELT para Extracto de Banco Hipotecario Dólares.
    Entidad: JOA (Cuentas Joaquín).
    """
    try:
        # 1. Calcular Hash SHA256 Legal
        sha256 = hashlib.sha256()
        with open(filepath, 'rb') as f:
            sha256.update(f.read())
        hash_hex = sha256.hexdigest()
        
        # 2. Generar representación Markdown Raw Completa (Fase 1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active # Toma la pestaña activa
        
        cuenta_alias = "CA U$D ...2646"
        
        raw_md_lines = []
        raw_md_lines.append("# EXTRACTO BANCARIO BANCO HIPOTECARIO - CAJA DE AHORRO DOLARES")
        raw_md_lines.append(f"**Cuenta:** {cuenta_alias}")
        raw_md_lines.append(f"**SHA256:** {hash_hex}\n")
        raw_md_lines.append("| Fecha | Movimiento / Descripción | Importe | Saldo Parcial |")
        raw_md_lines.append("| --- | --- | --- | --- |")
        
        movimientos_parsed = []
        
        # El archivo tiene filas de encabezado basura, los datos reales de movimientos empiezan más abajo (skiprows=4 en pandas)
        for r in range(6, ws.max_row + 1):
            fecha_val = ws.cell(r, 1).value
            desc_val = ws.cell(r, 2).value
            imp_val = ws.cell(r, 3).value
            saldo_val = ws.cell(r, 4).value
            
            if not fecha_val or not desc_val:
                continue
                
            clean_mov = " ".join(str(desc_val).split('\n')).strip()
            raw_md_lines.append(f"| {fecha_val} | {clean_mov} | {imp_val} | {saldo_val} |")
            
            def to_float(val, *, empty_value=0.0):
                if val in (None, ""): return empty_value
                try:
                    return float(str(val).replace('.', '').replace(',', '.'))
                except (TypeError, ValueError):
                    raise ValueError(f"Importe inválido en fila {r}: {val!r}")
                    
            imp_f = to_float(imp_val)
            saldo_f = to_float(saldo_val, empty_value=None)
            
            movimientos_parsed.append({
                'fecha': str(fecha_val),
                'descripcion': clean_mov,
                'importe': imp_f,
                'saldo': saldo_f,
                'cuenta': cuenta_alias,
                'banco': 'HIPOTECARIO'
            })
            
        raw_md_text = "\n".join(raw_md_lines)
        
        inserted, persist_info = storage_bancos.guardar_extracto_hipotecario(
            nombre_archivo=os.path.basename(filepath),
            hash_sha256=hash_hex,
            contenido_raw=raw_md_text,
            movimientos=movimientos_parsed,
            cuenta=cuenta_alias,
            moneda="USD",
            reconstruir=reconstruir,
        )
        staging_id = persist_info["staging_id"]
        if not inserted:
            return False, {"motivo": persist_info["motivo"], "id_insertado": staging_id}
        
        # 5. Correr el Verificador de Integridad Automático en Caliente
        try:
            from core_sistema.verificador_integridad import VerificadorIntegridadERP
            vi = VerificadorIntegridadERP()
            res_v = vi.verificar_y_reportar(filepath, staging_id)
            if res_v['status'] != 'OK':
                print(f"⚠️ [DISCREPANCIA DETECTADA EN INGESTA HIPOTECARIO USD] Staging ID: {staging_id} | Desvío: {res_v.get('diferencias')}")
        except Exception as ver_err:
            print(f"⚠️ Error al ejecutar verificador automático en caliente: {ver_err}")
            
        info = {
            "modulo": "bancos",
            "anio": datetime.now().strftime("%Y"),
            "mes": datetime.now().strftime("%m"),
            "entidad": "JOA",
            "db_table": "bancos_movimientos",
            "id_insertado": staging_id
        }
        return True, info
    except Exception as e:
        print(f"❌ Error en lector_hipotecario_usd: {e}")
        return False, {}
