import os
import shutil
import hashlib
import sqlite3
import openpyxl
from datetime import datetime

def procesar_archivo(filepath: str) -> tuple[bool, dict]:
    """
    Procesador ELT para Extracto de Banco Galicia (Caja de Ahorro / Cuenta Corriente).
    """
    try:
        # 1. Calcular Hash SHA256 Legal
        sha256 = hashlib.sha256()
        with open(filepath, 'rb') as f:
            sha256.update(f.read())
        hash_hex = sha256.hexdigest()
        
        # 2. Generar representación Markdown Raw Completa (Fase 1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['Cuentas']
        
        # Detectar si es Cuenta Corriente o Caja de Ahorros
        primer_celda = str(ws.cell(1, 1).value or "")
        nro_cuenta_celda = str(ws.cell(2, 1).value or "")
        
        is_cc = "CUENTA CORRIENTE" in primer_celda.upper() or "3762569" in nro_cuenta_celda
        
        tipo_cuenta_desc = "CUENTA CORRIENTE PESOS" if is_cc else "CAJA DE AHORRO PESOS"
        cuenta_alias = "CC$ ...3762569" if is_cc else "CA$ ...8342567"
        
        raw_md_lines = []
        raw_md_lines.append(f"# EXTRACTO BANCARIO GALICIA - {tipo_cuenta_desc}")
        raw_md_lines.append(f"**Cuenta:** {cuenta_alias}")
        raw_md_lines.append(f"**SHA256:** {hash_hex}\n")
        raw_md_lines.append("| Fecha | Movimiento / Descripción | Débito | Crédito | Saldo Parcial |")
        raw_md_lines.append("| --- | --- | --- | --- | --- |")
        
        movimientos_parsed = []
        
        for r in range(7, ws.max_row + 1):
            fecha_val = ws.cell(r, 1).value
            mov_val = str(ws.cell(r, 2).value or "").strip()
            deb_val = str(ws.cell(r, 3).value or "0,00").strip()
            cred_val = str(ws.cell(r, 4).value or "0,00").strip()
            saldo_val = str(ws.cell(r, 5).value or "0,00").strip()
            
            if not fecha_val or not mov_val:
                continue
                
            clean_mov = " ".join(mov_val.split('\n')).strip()
            raw_md_lines.append(f"| {fecha_val} | {clean_mov} | {deb_val} | {cred_val} | {saldo_val} |")
            
            def to_float(v_str):
                try:
                    v_clean = str(v_str).replace('.', '').replace(',', '.')
                    return float(v_clean)
                except:
                    return 0.0
                    
            deb_f = abs(to_float(deb_val))
            cred_f = abs(to_float(cred_val))
            neto_f = cred_f - deb_f if cred_f > 0 else -deb_f
            saldo_f = to_float(saldo_val)
            
            movimientos_parsed.append({
                'fecha': str(fecha_val),
                'descripcion': clean_mov,
                'importe': neto_f,
                'saldo': saldo_f,
                'cuenta': cuenta_alias,
                'banco': 'GALICIA'
            })
            
        raw_md_text = "\n".join(raw_md_lines)
        
        # 3. Guardar en core_staging_raw
        conn = sqlite3.connect('c:/Users/essao/Desktop/ERP FINAL/erp_nicoletti.db')
        conn.row_factory = sqlite3.Row
        
        cursor = conn.cursor()
        cursor.execute("DELETE FROM core_staging_raw WHERE hash_sha256 = ?", (hash_hex,))
        cursor.execute("""
            INSERT INTO core_staging_raw (
                nombre_archivo, hash_sha256, modulo, tipo_fuente, formato_raw, parser_version, contenido_raw, filas_leidas, estado, fecha_ingesta, fecha_procesado
            ) VALUES (
                ?, ?, 'bancos', 'EXCEL', 'MARKDOWN', 'v6.2.0', ?, ?, 'PROCESADO', datetime('now', 'localtime'), datetime('now', 'localtime')
            )
        """, (os.path.basename(filepath), hash_hex, raw_md_text, len(movimientos_parsed)))
        
        staging_id = cursor.lastrowid
        conn.commit()
        
        # 4. Insertar en bancos_movimientos
        conn.execute("DELETE FROM bancos_movimientos WHERE cuenta = ? AND raw_ingesta_id IN (SELECT id FROM core_staging_raw WHERE hash_sha256 = ?)", (cuenta_alias, hash_hex))
        
        for m in movimientos_parsed:
            cat = 'SIN_CATEGORIZAR'
            desc = m['descripcion'].upper()
            if 'AFIP' in desc or 'VEP' in desc or 'LEY 25413' in desc:
                cat = 'Impuestos'
            elif 'SEC' in desc or '30548970837' in desc:
                cat = 'Impuestos'
            elif 'SUELDO' in desc or 'HABERES' in desc:
                cat = 'Sueldo'
            elif 'INTERES' in desc:
                cat = 'Intereses'
            elif 'MERCADO LIBRE' in desc or 'MERCADOPAGO' in desc:
                cat = 'Compras'
            elif 'CALIM' in desc:
                cat = 'Servicios'
                
            conn.execute("""
                INSERT OR REPLACE INTO bancos_movimientos (fecha, cuenta, banco, descripcion, importe, saldo, categoria, raw_ingesta_id, entidad)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'LDK')
            """, (m['fecha'], m['cuenta'], m['banco'], m['descripcion'], m['importe'], m['saldo'], cat, staging_id))
            
        conn.commit()
        conn.close()
        
        # 5. Correr el Verificador de Integridad Automático en Caliente
        try:
            from core_sistema.verificador_integridad import VerificadorIntegridadERP
            vi = VerificadorIntegridadERP()
            res_v = vi.verificar_y_reportar(filepath, staging_id)
            if res_v['status'] != 'OK':
                print(f"⚠️ [DISCREPANCIA DETECTADA EN INGESTA GALICIA] Staging ID: {staging_id} | Desvío: {res_v.get('diferencias')}")
        except Exception as ver_err:
            print(f"⚠️ Error al ejecutar verificador automático en caliente: {ver_err}")
        
        info = {
            "modulo": "bancos",
            "anio": datetime.now().strftime("%Y"),
            "mes": datetime.now().strftime("%m"),
            "entidad": "LDK",
            "db_table": "bancos_movimientos",
            "id_insertado": staging_id
        }
        return True, info
    except Exception as e:
        print(f"❌ Error en lector_galicia: {e}")
        return False, {}
