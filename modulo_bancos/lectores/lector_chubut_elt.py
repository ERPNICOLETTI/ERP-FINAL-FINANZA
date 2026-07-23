import os
import shutil
import hashlib
import sqlite3
import openpyxl
from datetime import datetime

def procesar_archivo(filepath: str) -> tuple[bool, dict]:
    """
    Procesador ELT para Extractos de Banco Chubut (Caja de Ahorro / Cuenta Corriente).
    """
    try:
        # 1. Calcular Hash SHA256 Legal
        sha256 = hashlib.sha256()
        with open(filepath, 'rb') as f:
            sha256.update(f.read())
        hash_hex = sha256.hexdigest()
        
        # 2. Generar representación Markdown Raw Completa (Fase 1)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb['Historicos']
        
        # Detectar si es Cuenta Corriente o Caja de Ahorros
        nro_cuenta_celda = str(ws.cell(4, 3).value or "")
        is_cc = "CUENTA CORRIENTE" in nro_cuenta_celda.upper() or "00201" in nro_cuenta_celda
        
        tipo_cuenta_desc = "CUENTA CORRIENTE PESOS" if is_cc else "CAJA DE AHORRO PESOS"
        cuenta_alias = "CC$ ...00201" if is_cc else "CA$ ...00106"
        
        raw_md_lines = []
        raw_md_lines.append(f"# EXTRACTO BANCARIO BANCO CHUBUT - {tipo_cuenta_desc}")
        raw_md_lines.append(f"**Cuenta:** {cuenta_alias}")
        raw_md_lines.append(f"**SHA256:** {hash_hex}\n")
        raw_md_lines.append("| Fecha | Movimiento / Descripción | Código Movimiento | Importe |")
        raw_md_lines.append("| --- | --- | --- | --- |")
        
        movimientos_parsed = []
        
        for r in range(11, ws.max_row + 1):
            fecha_val = ws.cell(r, 2).value
            mov_val = str(ws.cell(r, 3).value or "").strip()
            cod_val = str(ws.cell(r, 4).value or "").strip()
            imp_val = str(ws.cell(r, 5).value or "0.00").strip()
            
            if not fecha_val or not mov_val:
                continue
                
            clean_mov = " ".join(mov_val.split('\n')).strip()
            raw_md_lines.append(f"| {fecha_val} | {clean_mov} | {cod_val} | {imp_val} |")
            
            def to_float(v_str):
                try:
                    return float(v_str)
                except:
                    return 0.0
                    
            imp_f = to_float(imp_val)
            
            movimientos_parsed.append({
                'fecha': str(fecha_val),
                'descripcion': f"{clean_mov} (Cod: {cod_val})",
                'importe': imp_f,
                'saldo': 0.0,
                'cuenta': cuenta_alias,
                'banco': 'CHUBUT'
            })
            
        raw_md_text = "\n".join(raw_md_lines)
        
        # 3. Guardar en core_staging_raw
        conn = sqlite3.connect('c:/Users/essao/Desktop/ERP FINAL/erp_nicoletti.db')
        conn.row_factory = sqlite3.Row
        
        cursor = conn.cursor()
        cursor.execute("DELETE FROM core_staging_raw WHERE hash_sha256 = ?", (hash_hex,))
        cursor.execute("""
            INSERT INTO core_staging_raw (
                nombre_archivo, hash_sha256, modulo, tipo_fuente, formato_raw, parser_version, contenido_raw, filas_leidas, estado, fecha_ingesta, fecha_procesado
            ) VALUES (
                ?, ?, 'bancos', 'EXCEL', 'MARKDOWN', 'v6.2.0', ?, ?, 'PROCESADO', datetime('now', 'localtime'), datetime('now', 'localtime')
            )
        """, (os.path.basename(filepath), hash_hex, raw_md_text, len(movimientos_parsed)))
        
        staging_id = cursor.lastrowid
        conn.commit()
        
        # 4. Insertar en bancos_movimientos
        conn.execute("DELETE FROM bancos_movimientos WHERE cuenta = ? AND raw_ingesta_id IN (SELECT id FROM core_staging_raw WHERE hash_sha256 = ?)", (cuenta_alias, hash_hex))
        
        for m in movimientos_parsed:
            cat = 'SIN_CATEGORIZAR'
            desc = m['descripcion'].upper()
            if 'AFIP' in desc or 'VEP' in desc or 'LEY 25413' in desc or 'IMPUESTO' in desc or 'SELLOS' in desc:
                cat = 'Impuestos'
            elif 'LIQUID' in desc or 'ACREDIT' in desc:
                cat = 'Ventas'
            elif 'SUELDO' in desc:
                cat = 'Sueldo'
            elif 'INTERES' in desc:
                cat = 'Intereses'
            elif 'MERCADO LIBRE' in desc or 'MERCADOPAGO' in desc:
                cat = 'Compras'
                
            conn.execute("""
                INSERT OR REPLACE INTO bancos_movimientos (fecha, cuenta, banco, descripcion, importe, saldo, categoria, raw_ingesta_id, entidad)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'LDK')
            """, (m['fecha'], m['cuenta'], m['banco'], m['descripcion'], m['importe'], m['saldo'], cat, staging_id))
            
        conn.commit()
        conn.close()
        
        info = {
            "modulo": "bancos",
            "anio": datetime.now().strftime("%Y"),
            "mes": datetime.now().strftime("%m"),
            "entidad": "LDK",
            "db_table": "bancos_movimientos",
            "id_insertado": staging_id
        }
        return True, info
    except Exception as e:
        print(f"❌ Error en lector_chubut_elt: {e}")
        return False, {}
